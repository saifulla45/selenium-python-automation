
import openpyxl
from openpyxl.styles import PatternFill

def get_row_count(file,sheet_name):
    workbook = openpyxl.load_workbook(filename=file)
    sheet = workbook[sheet_name]
    return (sheet.max_row)

def get_column_count(file,sheet_name):
    workbook = openpyxl.load_workbook(filename=file)
    sheet = workbook[sheet_name]
    return (sheet.max_column)

def read_data(file,sheet_name,row_num,column_num):
    workbook = openpyxl.load_workbook(filename=file)
    sheet = workbook[sheet_name]
    return (sheet.cell(row=row_num,column=column_num).value)

def write_data(file,sheet_name,row_num,column_num,data):
    workbook = openpyxl.load_workbook(filename=file)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_num,column=column_num).value = data
    workbook.save(filename=file)

def fill_green(file,sheet_name,row_num,column_num):
    workbook = openpyxl.load_workbook(filename=file)
    sheet = workbook[sheet_name]
    greenFill = PatternFill(start_color='00FF00',end_color='00FF00',patternType='solid')
    sheet.cell(row=row_num,column=column_num).fill = greenFill
    workbook.save(filename=file)
    
def fill_red(file,sheet_name,row_num,column_num):
    workbook = openpyxl.load_workbook(filename=file)
    sheet = workbook[sheet_name]
    redFill = PatternFill(start_color='FF0000',end_color='FF0000',patternType='solid')
    sheet.cell(row=row_num,column=column_num).fill = redFill
    workbook.save(filename=file)
    